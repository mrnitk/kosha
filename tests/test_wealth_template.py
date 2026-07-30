"""Tests for the net-worth Excel template: headers, import, and round-trip."""

from __future__ import annotations

from datetime import date

import pytest

from kosha import crypto, wealth, wealth_template as wt
from kosha.db import Database

FAST = crypto.Argon2Params(time_cost=1, memory_cost=8192, parallelism=1)
PW = "correct horse battery"


@pytest.fixture()
def db(tmp_path):
    d = Database(db_file=tmp_path / "kosha.db", salt_file=tmp_path / "kosha.salt")
    d.create(PW, params=FAST)
    yield d
    d.lock()


def _workbook(path, assets_rows, liab_rows=None, insurance_rows=None):
    """Build a template-shaped .xlsx from row lists (first row = header)."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = wt.ASSET_SHEET
    for row in assets_rows:
        ws.append(row)
    lw = wb.create_sheet(wt.LIABILITY_SHEET)
    for row in (liab_rows or [wt.LIABILITY_HEADERS]):
        lw.append(row)
    iw = wb.create_sheet(wt.INSURANCE_SHEET)
    for row in (insurance_rows or [wt.INSURANCE_HEADERS]):
        iw.append(row)
    wb.save(str(path))
    return path


# --- header parsing ----------------------------------------------------------

@pytest.mark.parametrize("text,expected", [
    ("Jun'26", date(2026, 6, 1)),
    ("Dec'23", date(2023, 12, 1)),
    ("Jun-26", date(2026, 6, 1)),
    ("Jun 2026", date(2026, 6, 1)),
    ("June 2026", date(2026, 6, 1)),
    ("2026-06", date(2026, 6, 1)),
    ("2026-06-15", date(2026, 6, 15)),
    ("15/06/2026", date(2026, 6, 15)),
])
def test_parse_period_header(text, expected):
    assert wt.parse_period_header(text) == expected


@pytest.mark.parametrize("text", ["Name", "Liquidity", "Invested", "", "Total Asset", "42"])
def test_non_date_headers_are_attributes(text):
    assert wt.parse_period_header(text) is None


@pytest.mark.parametrize("cell,expected", [
    ("1,23,456.78", 123456.78), ("₹5,27,000", 527000.0), ("", None),
    ("-", None), ("NA", None), ("(500)", -500.0), ("9.5%", 9.5),
])
def test_parse_number(cell, expected):
    assert wt.parse_number(cell) == expected


# --- importing ---------------------------------------------------------------

def test_import_creates_holdings_and_snapshots(tmp_path, db):
    """A sheet shaped like the user's spreadsheet imports in one go."""
    path = _workbook(tmp_path / "nw.xlsx", [
        ["Name", "Category", "Type", "Liquidity", "Owner", "Invested", "Dec'23", "Jun'26"],
        ["HDFC-Cash", "Bank", "Cash", "High", "Me", 0, 80000, 92000],
        ["SBI-FD/RD", "Bank", "Debt", "High", "Me", 500000, 230000, 500000],
        ["Mom-HDFC-Cash", "Bank", "Cash", "High", "Mom", 0, 10000, 547000],
    ])
    result = wt.import_wealth_template(db, path)
    assert result.assets_created == 3 and result.assets_updated == 0
    assert result.snapshots == ["2023-12-01", "2026-06-01"]
    assert result.values_recorded == 6
    assert not result.has_problems

    names = {a.name: a for a in wealth.list_assets(db)}
    assert names["Mom-HDFC-Cash"].owner == "Mom"
    assert names["SBI-FD/RD"].invested == 500000
    series = wealth.networth_series(db)
    assert [p.as_of for p in series] == ["2023-12-01", "2026-06-01"]
    assert series[0].assets == 80000 + 230000 + 10000
    assert series[1].assets == 92000 + 500000 + 547000


def test_import_is_idempotent_and_updates_in_place(tmp_path, db):
    rows = [
        ["Name", "Category", "Type", "Liquidity", "Owner", "Invested", "Jun'26"],
        ["HDFC-Cash", "Bank", "Cash", "High", "Me", 0, 92000],
    ]
    path = _workbook(tmp_path / "nw.xlsx", rows)
    wt.import_wealth_template(db, path)
    # Same names, changed value -> updates rather than duplicating.
    rows[1] = ["hdfc-cash", "Bank", "Cash", "High", "Me", 0, 99000]   # different case
    path2 = _workbook(tmp_path / "nw2.xlsx", rows)
    result = wt.import_wealth_template(db, path2)
    assert result.assets_created == 0 and result.assets_updated == 1
    assert len(wealth.list_assets(db)) == 1
    assert wealth.current_networth(db).assets == 99000


def test_import_liabilities_with_emi_and_outstanding(tmp_path, db):
    path = _workbook(
        tmp_path / "nw.xlsx",
        [["Name", "Category", "Type", "Liquidity", "Owner", "Invested", "Jun'26"],
         ["HDFC-Cash", "Bank", "Cash", "High", "Me", 0, 900000]],
        liab_rows=[
            ["Name", "Kind", "Owner", "Principal", "Interest rate", "EMI",
             "Start date", "End date", "Jun'26"],
            ["Car loan - HDFC", "Car loan", "Me", 800000, 9.5, 16000,
             "2026-07-01", "2031-07-01", 720000],
        ])
    result = wt.import_wealth_template(db, path)
    assert result.liabilities_created == 1
    liab = wealth.list_liabilities(db)[0]
    assert liab.emi_amount == 16000 and liab.interest_rate == 9.5
    assert liab.start_date == "2026-07-01"
    point = wealth.current_networth(db)
    assert point.assets == 900000 and point.liabilities == 720000
    assert point.net_worth == 180000
    assert wealth.monthly_obligations(db) == 16000


def test_import_insurance_not_counted(tmp_path, db):
    path = _workbook(
        tmp_path / "nw.xlsx",
        [["Name", "Category", "Type", "Liquidity", "Owner", "Invested", "Jun'26"],
         ["HDFC-Cash", "Bank", "Cash", "High", "Me", 0, 92000]],
        insurance_rows=[["Name", "Kind", "Owner", "Premium per year", "Coverage"],
                        ["Medical - Self", "Medical", "Me", 16000, 2000000]])
    result = wt.import_wealth_template(db, path)
    assert result.insurance_created == 1
    assert wealth.current_networth(db).assets == 92000     # insurance excluded
    assert wealth.insurance_summary(db) == (16000, 2000000)


def test_blank_cells_are_skipped_not_zeroed(tmp_path, db):
    path = _workbook(tmp_path / "nw.xlsx", [
        ["Name", "Category", "Type", "Liquidity", "Owner", "Invested", "Dec'23", "Jun'26"],
        ["Gold", "Other", "Equity", "Medium", "Me", 0, 50000, None],   # not updated in Jun
    ])
    wt.import_wealth_template(db, path)
    # December value recorded; June has no explicit row, so it carries forward.
    assert wealth.snapshot_dates(db) == ["2023-12-01"]
    assert wealth.current_networth(db).assets == 50000


def test_negative_and_duplicate_rows_are_reported(tmp_path, db):
    path = _workbook(tmp_path / "nw.xlsx", [
        ["Name", "Category", "Type", "Liquidity", "Owner", "Invested", "Jun'26"],
        ["Cash", "Bank", "Cash", "High", "Me", 0, -500],       # negative -> skipped
        ["Cash", "Bank", "Cash", "High", "Me", 0, 100],        # duplicate -> skipped
        ["FD", "Bank", "Debt", "High", "Me", 0, 200],
    ])
    result = wt.import_wealth_template(db, path)
    assert result.has_problems and len(result.problems) == 2
    joined = " ".join(result.problems).lower()
    assert "negative" in joined and "duplicate" in joined
    assert wealth.current_networth(db).assets == 200          # only the clean row


def test_missing_sheets_raise_template_error(tmp_path, db):
    import openpyxl
    wb = openpyxl.Workbook()
    wb.active.append(["nothing", "useful"])
    path = tmp_path / "bad.xlsx"
    wb.save(str(path))
    with pytest.raises(wt.WealthTemplateError):
        wt.import_wealth_template(db, path)


def test_alias_headers_accepted(tmp_path, db):
    """The user's own column names ('Where', 'Whose') work too."""
    path = _workbook(tmp_path / "nw.xlsx", [
        ["Where", "Category", "Type", "Liquidity", "Whose", "Invested", "Jun'26"],
        ["HDFC-Cash", "Bank", "Cash", "High", "Mom", 0, 92000],
    ])
    result = wt.import_wealth_template(db, path)
    assert result.assets_created == 1
    assert wealth.list_assets(db)[0].owner == "Mom"


# --- template generation / round trip ----------------------------------------

def test_write_blank_template_is_importable(tmp_path, db):
    out = tmp_path / "blank.xlsx"
    wt.write_wealth_template(out)
    assert out.exists()
    import openpyxl
    wb = openpyxl.load_workbook(out)
    assert {wt.ASSET_SHEET, wt.LIABILITY_SHEET, wt.INSURANCE_SHEET} <= set(wb.sheetnames)
    # The sample rows carry no date columns, so there's nothing to import yet.
    with pytest.raises(wt.WealthTemplateError):
        wt.import_wealth_template(db, out)


def test_export_round_trips_existing_data(tmp_path, db):
    cash = wealth.add_asset(db, "HDFC-Cash", "Bank", "Cash", "High")
    lid = wealth.add_liability(db, "Car loan", "Car loan", emi_amount=16000)
    wealth.record_snapshot(db, date(2025, 12, 1), {cash: 81000}, {lid: 800000})
    wealth.record_snapshot(db, date(2026, 6, 1), {cash: 92000}, {lid: 720000})
    before = wealth.networth_series(db)

    out = tmp_path / "export.xlsx"
    wt.write_wealth_template(out, db)

    # Import the export into a fresh vault: the series must match.
    other = Database(db_file=tmp_path / "o.db", salt_file=tmp_path / "o.salt")
    other.create(PW, params=FAST)
    result = wt.import_wealth_template(other, out)
    assert result.assets_created == 1 and result.liabilities_created == 1
    after = wealth.networth_series(other)
    assert [(p.as_of, p.assets, p.liabilities, p.net_worth) for p in after] == \
           [(p.as_of, p.assets, p.liabilities, p.net_worth) for p in before]
    other.lock()


def test_read_table_sheet_selection(tmp_path):
    """tabular.read_table can pick a sheet by name, case-insensitively."""
    from kosha import tabular
    import openpyxl
    wb = openpyxl.Workbook()
    wb.active.title = "First"
    wb.active.append(["a"])
    second = wb.create_sheet("Second")
    second.append(["b"])
    path = tmp_path / "two.xlsx"
    wb.save(str(path))
    assert tabular.read_table(path)[0] == ["a"]                 # default: first
    assert tabular.read_table(path, sheet="second")[0] == ["b"]  # by name
    with pytest.raises(KeyError):
        tabular.read_table(path, sheet="missing")
