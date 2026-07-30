"""Excel import/export for net-worth history.

The template mirrors how the user already tracks this in a spreadsheet: one row
per holding, fixed attribute columns, then **one column per snapshot date**.

    Assets sheet
        Name | Category | Type | Liquidity | Owner | Invested | Dec'23 | Apr'24 | ...

    Liabilities sheet
        Name | Kind | Owner | Principal | Interest rate | EMI | Start | End | Dec'23 | ...

Any header that parses as a date (``Jun'26``, ``Jun-2026``, ``2026-06``,
``2026-06-01`` …) becomes a snapshot column; everything else is treated as an
attribute. Month-only headers are stored as the 1st of that month.

Importing is idempotent: holdings are matched by name (case-insensitively), so
re-importing an updated sheet edits in place instead of duplicating rows.
"""

from __future__ import annotations

import re
from datetime import date, datetime
from typing import Optional

from . import tabular, wealth
from .db import Database

ASSET_SHEET = "Assets"
LIABILITY_SHEET = "Liabilities"
INSURANCE_SHEET = "Insurance"

ASSET_HEADERS = ["Name", "Category", "Type", "Liquidity", "Owner", "Invested"]
LIABILITY_HEADERS = ["Name", "Kind", "Owner", "Principal", "Interest rate", "EMI",
                     "Start date", "End date"]
INSURANCE_HEADERS = ["Name", "Kind", "Owner", "Premium per year", "Coverage"]

# Attribute header -> canonical field, matched case-insensitively.
_ASSET_FIELDS = {
    "name": "name", "where": "name", "holding": "name",
    "category": "category", "group": "category",
    "type": "asset_type", "asset type": "asset_type",
    "liquidity": "liquidity",
    "owner": "owner", "whose": "owner",
    "invested": "invested", "cost": "invested", "invested amount": "invested",
}
_LIABILITY_FIELDS = {
    "name": "name", "loan": "name",
    "kind": "kind", "type": "kind",
    "owner": "owner",
    "principal": "principal", "sanctioned": "principal",
    "interest rate": "interest_rate", "rate": "interest_rate", "interest": "interest_rate",
    "emi": "emi_amount", "emi amount": "emi_amount", "instalment": "emi_amount",
    "start date": "start_date", "start": "start_date",
    "end date": "end_date", "end": "end_date",
}
_INSURANCE_FIELDS = {
    "name": "name", "policy": "name",
    "kind": "kind", "type": "kind",
    "owner": "owner",
    "premium per year": "premium_per_year", "premium": "premium_per_year",
    "coverage": "coverage", "cover": "coverage",
}

_MONTHS = {m.lower(): i for i, m in enumerate(
    ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"], 1)}
_FULL_MONTHS = {m.lower(): i for i, m in enumerate(
    ["January", "February", "March", "April", "May", "June", "July", "August",
     "September", "October", "November", "December"], 1)}


class WealthTemplateError(ValueError):
    """Raised when a file doesn't look like the net-worth template."""


# --- header / value parsing --------------------------------------------------

def parse_period_header(text: str) -> Optional[date]:
    """Parse a snapshot-column header into a date, or None if it isn't one.

    Accepts ``Jun'26`` / ``Jun-26`` / ``Jun 2026`` / ``June 2026`` / ``2026-06``
    / ``2026-06-01`` / ``01/06/2026``. Month-only forms become the 1st.
    """
    s = (text or "").strip()
    if not s:
        return None

    # Full ISO or common numeric dates first.
    for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    # Year-month, e.g. 2026-06
    m = re.fullmatch(r"(\d{4})[-/](\d{1,2})", s)
    if m:
        year, month = int(m.group(1)), int(m.group(2))
        if 1 <= month <= 12:
            return date(year, month, 1)
    # Month + year in words: Jun'26, Jun-26, Jun 2026, June 2026
    m = re.fullmatch(r"([A-Za-z]{3,9})\s*['\-/ ]\s*(\d{2}|\d{4})", s)
    if m:
        name, year_text = m.group(1).lower(), m.group(2)
        month = _MONTHS.get(name[:3]) if name[:3] in _MONTHS else None
        if name in _FULL_MONTHS:
            month = _FULL_MONTHS[name]
        if month:
            year = int(year_text)
            if year < 100:                      # two-digit year -> 2000s
                year += 2000
            return date(year, month, 1)
    return None


def parse_number(cell: str) -> Optional[float]:
    """Parse a money/number cell, or None when blank or unusable.

    None means "no value recorded" (so the cell is skipped), which is different
    from 0.0 meaning "this holding was empty on that date".
    """
    s = (cell or "").strip()
    if not s or s in {"-", "–", "—", "NA", "na", "N/A", "n/a"}:
        return None
    neg = s.startswith("(") and s.endswith(")")
    for junk in ("₹", ",", " ", "Rs.", "Rs", "INR"):
        s = s.replace(junk, "")
    s = s.strip("()").rstrip("%")
    try:
        value = float(s)
    except ValueError:
        return None
    return -value if neg else value


def _norm(text: str) -> str:
    return " ".join(str(text or "").strip().lower().replace(".", " ").replace("_", " ").split())


def _find_header_row(rows: list[list[str]], fields: dict[str, str],
                     require_dates: bool = True):
    """Locate the header row: the first row with a Name column (plus a date column
    when ``require_dates``, which insurance sheets don't have).

    Returns ``(index, {field: col}, {col: date})``.
    """
    for i, row in enumerate(rows[:30]):
        attrs: dict[str, int] = {}
        dates: dict[int, date] = {}
        for c, cell in enumerate(row):
            as_date = parse_period_header(cell)
            if as_date is not None:
                dates[c] = as_date
                continue
            field = fields.get(_norm(cell))
            if field and field not in attrs:
                attrs[field] = c
        if "name" in attrs and (dates or not require_dates):
            return i, attrs, dates
    return None


def _cell(row: list[str], idx: Optional[int]) -> str:
    if idx is None or idx < 0 or idx >= len(row):
        return ""
    return row[idx]


# --- reading -----------------------------------------------------------------

def _read_sheet(path, sheet: str, fields: dict[str, str], require_dates: bool = True):
    """Read one sheet into (attribute dicts, {name: {date: value}}, problems)."""
    try:
        rows = tabular.read_table(path, sheet=sheet)
    except (ValueError, KeyError):
        return [], {}, []                      # sheet absent — nothing to import
    found = _find_header_row(rows, fields, require_dates)
    if found is None:
        return [], {}, []
    hidx, attrs, date_cols = found

    records: list[dict] = []
    values: dict[str, dict[date, float]] = {}
    problems: list[str] = []
    seen: set[str] = set()
    for idx in range(hidx + 1, len(rows)):
        row = rows[idx]
        sheet_row = idx + 1                     # 1-based, as Excel shows it
        name = _cell(row, attrs.get("name")).strip()
        if not name:
            continue
        key = name.lower()
        if key in seen:
            problems.append(f"{sheet} row {sheet_row}: duplicate holding '{name}' — skipped")
            continue
        seen.add(key)

        record = {"name": name}
        for field, col in attrs.items():
            if field == "name":
                continue
            raw = _cell(row, col).strip()
            if not raw:
                continue
            if field in ("invested", "principal", "emi_amount", "interest_rate",
                         "premium_per_year", "coverage"):
                number = parse_number(raw)
                if number is None:
                    continue
                if number < 0:
                    problems.append(
                        f"{sheet} row {sheet_row}: negative {field.replace('_', ' ')} — ignored")
                    continue
                record[field] = number
            else:
                record[field] = raw
        records.append(record)

        per_date: dict[date, float] = {}
        for col, as_of in date_cols.items():
            number = parse_number(_cell(row, col))
            if number is None:
                continue
            if number < 0:
                problems.append(
                    f"{sheet} row {sheet_row}: negative value for {as_of.isoformat()} — skipped")
                continue
            per_date[as_of] = number
        values[name] = per_date
    return records, values, problems


def read_wealth_template(path):
    """Read the whole workbook.

    Returns ``(assets, asset_values, liabilities, liability_values, insurance,
    problems)``. Raises :class:`WealthTemplateError` if no usable sheet is found.
    """
    assets, asset_values, problems = _read_sheet(path, ASSET_SHEET, _ASSET_FIELDS)
    liabs, liab_values, liab_problems = _read_sheet(path, LIABILITY_SHEET, _LIABILITY_FIELDS)
    insurance, _iv, ins_problems = _read_sheet(
        path, INSURANCE_SHEET, _INSURANCE_FIELDS, require_dates=False)
    problems = problems + liab_problems + ins_problems
    if not assets and not liabs:
        raise WealthTemplateError(
            "No net-worth data found. The workbook needs an 'Assets' (or "
            "'Liabilities') sheet with a Name column and at least one date "
            "column such as Jun'26 — download the template from "
            "File ▸ Download net-worth template.")
    return assets, asset_values, liabs, liab_values, insurance, problems


# --- importing ---------------------------------------------------------------

class WealthImportResult:
    """What an import did, for the summary dialog."""

    def __init__(self):
        self.assets_created = 0
        self.assets_updated = 0
        self.liabilities_created = 0
        self.liabilities_updated = 0
        self.insurance_created = 0
        self.snapshots: list[str] = []
        self.values_recorded = 0
        self.problems: list[str] = []

    @property
    def has_problems(self) -> bool:
        return bool(self.problems)

    def summary(self) -> str:
        lines = [
            f"Assets: {self.assets_created} added, {self.assets_updated} updated.",
            f"Liabilities: {self.liabilities_created} added, "
            f"{self.liabilities_updated} updated.",
        ]
        if self.insurance_created:
            lines.append(f"Insurance: {self.insurance_created} added.")
        lines.append(f"Recorded {self.values_recorded} value(s) across "
                     f"{len(self.snapshots)} snapshot date(s).")
        if self.snapshots:
            shown = ", ".join(self.snapshots[:8])
            more = f" … (+{len(self.snapshots) - 8})" if len(self.snapshots) > 8 else ""
            lines.append(f"Dates: {shown}{more}")
        if self.problems:
            lines.append("")
            lines.append(f"⚠ {len(self.problems)} issue(s) were skipped:")
            lines.extend(f"  • {p}" for p in self.problems[:12])
            if len(self.problems) > 12:
                lines.append(f"  … and {len(self.problems) - 12} more")
        return "\n".join(lines)


def import_wealth_template(db: Database, path) -> WealthImportResult:
    """Create/update holdings from the workbook and record every snapshot column."""
    (assets, asset_values, liabs, liab_values,
     insurance, problems) = read_wealth_template(path)
    result = WealthImportResult()
    result.problems = list(problems)

    existing_assets = {a.name.lower(): a for a in wealth.list_assets(db)}
    asset_ids: dict[str, int] = {}
    for record in assets:
        name = record["name"]
        current = existing_assets.get(name.lower())
        fields = {
            "category": record.get("category", "Other"),
            "asset_type": record.get("asset_type", "Cash"),
            "liquidity": record.get("liquidity", "High"),
            "owner": record.get("owner", "Me"),
            "invested": record.get("invested", 0.0),
        }
        if current:
            wealth.update_asset(db, current.id, **fields)
            asset_ids[name] = current.id
            result.assets_updated += 1
        else:
            asset_ids[name] = wealth.add_asset(db, name, **fields)
            result.assets_created += 1

    existing_liabs = {l.name.lower(): l for l in wealth.list_liabilities(db)}
    liab_ids: dict[str, int] = {}
    for record in liabs:
        name = record["name"]
        current = existing_liabs.get(name.lower())
        fields = {
            "kind": record.get("kind", "Other"),
            "owner": record.get("owner", "Me"),
            "principal": record.get("principal", 0.0),
            "interest_rate": record.get("interest_rate"),
            "emi_amount": record.get("emi_amount", 0.0),
        }
        for key in ("start_date", "end_date"):
            parsed = parse_period_header(record.get(key, ""))
            if parsed:
                fields[key] = parsed.isoformat()
        if current:
            wealth.update_liability(db, current.id, **fields)
            liab_ids[name] = current.id
            result.liabilities_updated += 1
        else:
            liab_ids[name] = wealth.add_liability(db, name, **fields)
            result.liabilities_created += 1

    existing_ins = {p.name.lower() for p in wealth.list_insurance(db)}
    for record in insurance:
        if record["name"].lower() in existing_ins:
            continue                            # don't duplicate policies
        wealth.add_insurance(
            db, record["name"], record.get("kind"),
            premium_per_year=record.get("premium_per_year", 0.0),
            coverage=record.get("coverage", 0.0), owner=record.get("owner", "Me"))
        result.insurance_created += 1

    # Group every value by date so each date is written as one snapshot.
    by_date: dict[date, tuple[dict[int, float], dict[int, float]]] = {}
    for name, per_date in asset_values.items():
        holding_id = asset_ids.get(name)
        if holding_id is None:
            continue
        for as_of, value in per_date.items():
            by_date.setdefault(as_of, ({}, {}))[0][holding_id] = value
    for name, per_date in liab_values.items():
        holding_id = liab_ids.get(name)
        if holding_id is None:
            continue
        for as_of, value in per_date.items():
            by_date.setdefault(as_of, ({}, {}))[1][holding_id] = value

    for as_of in sorted(by_date):
        asset_map, liab_map = by_date[as_of]
        wealth.record_snapshot(db, as_of, asset_map, liab_map)
        result.snapshots.append(as_of.isoformat())
        result.values_recorded += len(asset_map) + len(liab_map)
    return result


# --- template generation -----------------------------------------------------

def write_wealth_template(path, db: Optional[Database] = None) -> None:
    """Write a blank (or pre-filled) net-worth template workbook.

    When ``db`` is given, existing holdings and snapshot columns are exported, so
    the file doubles as a backup/round-trip of what's already recorded.
    """
    import openpyxl
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    bold = Font(bold=True)

    dates = wealth.snapshot_dates(db) if db else []
    date_headers = [d[:7] for d in dates]        # 'YYYY-MM' reads well as a column

    ws = wb.active
    ws.title = ASSET_SHEET
    ws.append(ASSET_HEADERS + date_headers)
    for cell in ws[1]:
        cell.font = bold
    if db:
        for asset in wealth.list_assets(db):
            recorded = dict(db.connection.execute(
                "SELECT as_of, value FROM asset_valuations WHERE asset_id=?",
                (asset.id,)).fetchall())
            ws.append([asset.name, asset.category, asset.asset_type, asset.liquidity,
                       asset.owner, asset.invested] + [recorded.get(d) for d in dates])
    else:
        ws.append(["HDFC-Cash", "Bank", "Cash", "High", "Me", 0])
        ws.append(["Mine (Non Tax saving)", "Mutual funds", "Equity", "Medium", "Me", 600000])
    _widths(ws, [30, 16, 12, 12, 10, 14] + [14] * len(date_headers), get_column_letter)
    ws.freeze_panes = "B2"

    lw = wb.create_sheet(LIABILITY_SHEET)
    lw.append(LIABILITY_HEADERS + date_headers)
    for cell in lw[1]:
        cell.font = bold
    if db:
        for liab in wealth.list_liabilities(db):
            recorded = dict(db.connection.execute(
                "SELECT as_of, outstanding FROM liability_valuations WHERE liability_id=?",
                (liab.id,)).fetchall())
            lw.append([liab.name, liab.kind, liab.owner, liab.principal,
                       liab.interest_rate, liab.emi_amount, liab.start_date,
                       liab.end_date] + [recorded.get(d) for d in dates])
    else:
        lw.append(["Car loan - HDFC", "Car loan", "Me", 800000, 9.5, 16000, "", ""])
    _widths(lw, [30, 14, 10, 14, 12, 12, 12, 12] + [14] * len(date_headers), get_column_letter)
    lw.freeze_panes = "B2"

    iw = wb.create_sheet(INSURANCE_SHEET)
    iw.append(INSURANCE_HEADERS)
    for cell in iw[1]:
        cell.font = bold
    if db:
        for policy in wealth.list_insurance(db):
            iw.append([policy.name, policy.kind, policy.owner,
                       policy.premium_per_year, policy.coverage])
    else:
        iw.append(["Medical - Self - HDFC ergo", "Medical", "Me", 16000, 2000000])
    _widths(iw, [34, 12, 10, 18, 16], get_column_letter)

    info = wb.create_sheet("Instructions")
    for line in [
        ["Kosha net-worth template — how to use"],
        [""],
        ["One row per holding. Attribute columns first, then ONE COLUMN PER DATE."],
        [""],
        ["Assets sheet", "Name, Category, Type, Liquidity, Owner, Invested, then date columns"],
        ["Liabilities sheet", "Name, Kind, Owner, Principal, Interest rate, EMI, Start, End, then dates"],
        ["Insurance sheet", "Name, Kind, Owner, Premium per year, Coverage (never counted in net worth)"],
        [""],
        ["Date column headers", "Jun'26  |  Jun-2026  |  June 2026  |  2026-06  |  2026-06-01"],
        ["", "Month-only headers are stored as the 1st of that month."],
        ["Values", "The amount held (assets) or still owed (liabilities) on that date."],
        ["", "Leave a cell blank if you didn't record that holding on that date."],
        ["Type", "Cash / Debt / Equity / Hybrid"],
        ["Liquidity", "High / Medium / Low / Lowest"],
        ["Kind (liabilities)", "Home loan / Car loan / Personal loan / Credit card / Other"],
        [""],
        ["Re-importing is safe", "Holdings are matched by name, so an updated sheet edits in place."],
        ["Net worth", "= total assets − total liabilities on each date."],
    ]:
        info.append(line)
    info["A1"].font = Font(bold=True, size=13)
    _widths(info, [26, 74], get_column_letter)

    wb.save(str(path))


def _widths(ws, widths, get_column_letter) -> None:
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
