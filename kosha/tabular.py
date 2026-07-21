"""Read any supported statement file into a matrix of string cells.

Backs the generic column-mapping importer: CSV/TSV, .xlsx/.xlsm (openpyxl) and
legacy .xls (xlrd) all reduce to ``list[list[str]]`` so the mapping UI and the
generic parser don't care about the file format. Cells are stringified (dates to
ISO ``YYYY-MM-DD``) so the same mapping works across formats.
"""

from __future__ import annotations

import csv
from datetime import date, datetime
from pathlib import Path

CSV_SUFFIXES = {".csv", ".tsv", ".txt"}
XLSX_SUFFIXES = {".xlsx", ".xlsm"}
XLS_SUFFIXES = {".xls"}
SUPPORTED_SUFFIXES = CSV_SUFFIXES | XLSX_SUFFIXES | XLS_SUFFIXES


def _stringify(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        # Drop a midnight time component so date columns read cleanly.
        return value.strftime("%Y-%m-%d" if value.time() == datetime.min.time() else "%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _read_csv(path: Path) -> list[list[str]]:
    # utf-8-sig strips a BOM if present; fall back to latin-1 for odd exports.
    for encoding in ("utf-8-sig", "latin-1"):
        try:
            with open(path, newline="", encoding=encoding) as fh:
                sample = fh.read(4096)
                fh.seek(0)
                delimiter = "\t" if path.suffix.lower() == ".tsv" else _sniff_delimiter(sample)
                return [[_stringify(c) for c in row] for row in csv.reader(fh, delimiter=delimiter)]
        except UnicodeDecodeError:
            continue
    raise ValueError(f"could not decode {path.name} as text")


def _sniff_delimiter(sample: str) -> str:
    try:
        return csv.Sniffer().sniff(sample, delimiters=",;\t|").delimiter
    except csv.Error:
        return ","


def _read_xlsx(path: Path) -> list[list[str]]:
    import openpyxl
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    try:
        sheet = wb[wb.sheetnames[0]]
        return [[_stringify(c) for c in row] for row in sheet.iter_rows(values_only=True)]
    finally:
        wb.close()


def _read_xls(path: Path) -> list[list[str]]:
    import xlrd
    book = xlrd.open_workbook(str(path))
    sheet = book.sheet_by_index(0)
    rows: list[list[str]] = []
    for r in range(sheet.nrows):
        row = []
        for c in range(sheet.ncols):
            cell = sheet.cell(r, c)
            # xlrd stores dates as floats with a flag; convert them to ISO.
            if cell.ctype == xlrd.XL_CELL_DATE:
                y, mo, d, *_ = xlrd.xldate_as_tuple(cell.value, book.datemode)
                row.append(f"{y:04d}-{mo:02d}-{d:02d}")
            else:
                row.append(_stringify(cell.value))
        rows.append(row)
    return rows


def read_table(path) -> list[list[str]]:
    """Read ``path`` into rows of string cells. Rows are ragged (as in the file)."""
    path = Path(path)
    suffix = path.suffix.lower()
    if suffix in CSV_SUFFIXES:
        return _read_csv(path)
    if suffix in XLSX_SUFFIXES:
        return _read_xlsx(path)
    if suffix in XLS_SUFFIXES:
        return _read_xls(path)
    raise ValueError(f"unsupported file type {suffix!r} (use CSV, XLSX or XLS)")


def column_count(rows: list[list[str]]) -> int:
    return max((len(r) for r in rows), default=0)


def guess_header_row(rows: list[list[str]], scan: int = 25) -> int:
    """Best-guess index of the header row: the early row with the most non-empty,
    mostly-non-numeric cells. Falls back to 0. The UI lets the user override."""
    best_idx, best_score = 0, -1.0
    for i, row in enumerate(rows[:scan]):
        cells = [c for c in row if c.strip()]
        if len(cells) < 2:
            continue
        wordy = sum(1 for c in cells if not _looks_numeric(c))
        score = len(cells) + wordy
        if score > best_score:
            best_idx, best_score = i, score
    return best_idx


def _looks_numeric(cell: str) -> bool:
    s = cell.replace(",", "").replace("₹", "").replace("(", "-").replace(")", "").strip()
    try:
        float(s)
        return True
    except ValueError:
        return False
