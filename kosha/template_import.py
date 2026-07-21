"""Standard-template import: one fixed spreadsheet layout for any bank.

Instead of teaching Kosha every bank's export format, the user pours their data
into one standard template and uploads it. The template columns are:

    Date | Transaction Remarks | Debit | Credit | Source | Account Type

``Source`` (e.g. "HDFC Bank", "SBI Savings", "Axis Card") is what shows up as the
transaction's source, so one file can hold many accounts. ``Account Type`` is
optional (inferred from the source name when blank). Only Date, Remarks and one
of Debit/Credit are required per row.

Reading is tolerant: headers match case-insensitively with common aliases, the
header row is found by content (so title/metadata rows above it are fine), and
rows without a parseable date are skipped (blank lines, subtotals, footers).
"""

from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
from typing import Optional

from . import tabular
from .parsers.base import RawTransaction

# The canonical template columns, in order.
TEMPLATE_HEADERS = ["Date", "Transaction Remarks", "Debit", "Credit", "Source", "Account Type"]

# field key -> accepted header spellings (lower-cased, punctuation-trimmed)
_ALIASES: dict[str, set[str]] = {
    "date": {"date", "txn date", "transaction date", "value date", "posting date"},
    "remarks": {"transaction remarks", "remarks", "description", "narration",
                "particulars", "details", "transaction details"},
    "debit": {"debit", "withdrawal", "withdrawal amt", "withdrawal amount",
              "dr", "debit amount", "amount debit", "paid out"},
    "credit": {"credit", "deposit", "deposit amt", "deposit amount",
               "cr", "credit amount", "amount credit", "paid in"},
    "source": {"source", "account", "account name", "bank", "bank name"},
    "account_type": {"account type", "type"},
}

_FALLBACK_DATE_FORMATS = (
    "%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%d/%m/%Y", "%d-%m-%Y", "%d/%m/%y",
    "%d-%m-%y", "%m/%d/%Y", "%m/%d/%y", "%d %b %Y", "%d-%b-%Y", "%d %B %Y",
    "%Y/%m/%d",
)


class TemplateError(ValueError):
    """Raised when a file doesn't look like the standard template."""


# --- value parsing -----------------------------------------------------------

def parse_amount(cell: str) -> float:
    """Money cell -> float: handles ₹, commas (incl. Indian), parenthesised negatives."""
    s = (cell or "").strip()
    if not s:
        return 0.0
    neg = s.startswith("(") and s.endswith(")")
    for junk in ("₹", ",", " ", "Rs.", "INR", "Dr", "Cr", "DR", "CR"):
        s = s.replace(junk, "")
    s = s.strip("()")
    try:
        val = float(s)
    except ValueError:
        return 0.0
    return -val if neg else val


def parse_date(cell: str) -> Optional[date]:
    """Parse a date cell, trying several common formats (and ISO from Excel cells)."""
    s = (cell or "").strip()
    if not s:
        return None
    for fmt in _FALLBACK_DATE_FORMATS:
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


# --- header detection --------------------------------------------------------

def _norm_header(cell: str) -> str:
    return " ".join(str(cell).strip().lower().replace(".", " ").replace("_", " ").split())


def _header_field(cell: str) -> Optional[str]:
    norm = _norm_header(cell)
    for field, names in _ALIASES.items():
        if norm in names:
            return field
    return None


def _find_header_row(rows: list[list[str]]):
    """Return (row_index, {field: col_index}). Raises TemplateError if not found."""
    for i, row in enumerate(rows[:25]):
        cols: dict[str, int] = {}
        for c, cell in enumerate(row):
            field = _header_field(cell)
            if field and field not in cols:
                cols[field] = c
        if "date" in cols and "remarks" in cols and ("debit" in cols or "credit" in cols):
            return i, cols
    raise TemplateError(
        "Couldn't find the template headers. The first row should contain "
        "Date, Transaction Remarks, Debit and Credit — download the blank "
        "template from File ▸ Download template and fill it in."
    )


def _cell(row: list[str], idx: Optional[int]) -> str:
    if idx is None or idx < 0 or idx >= len(row):
        return ""
    return row[idx]


def _account_type(type_text: str, source: str) -> str:
    hint = f"{type_text} {source}".lower()
    if "credit" in hint or "card" in hint:
        return "credit_card"
    return "bank"


def read_template(path) -> list[tuple[RawTransaction, str, str]]:
    """Read a filled template into (RawTransaction, source_name, account_type) rows.

    Rows without a parseable date or a debit/credit amount are skipped.
    """
    rows = tabular.read_table(path)
    hidx, cols = _find_header_row(rows)
    out: list[tuple[RawTransaction, str, str]] = []
    for row in rows[hidx + 1:]:
        txn_date = parse_date(_cell(row, cols["date"]))
        if txn_date is None:
            continue
        debit = parse_amount(_cell(row, cols.get("debit")))
        credit = parse_amount(_cell(row, cols.get("credit")))
        if debit > 0:
            amount, direction = debit, "debit"
        elif credit > 0:
            amount, direction = credit, "credit"
        else:
            continue
        remarks = _cell(row, cols.get("remarks")).strip() or "(no description)"
        source = _cell(row, cols.get("source")).strip() or "Imported"
        atype = _account_type(_cell(row, cols.get("account_type")), source)
        out.append((RawTransaction(txn_date, remarks, amount, direction), source, atype))
    return out


# --- blank template generation -----------------------------------------------

def write_template(path) -> None:
    """Write a blank .xlsx template: a Transactions sheet + an Instructions sheet."""
    import openpyxl
    from openpyxl.styles import Font

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transactions"
    ws.append(TEMPLATE_HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    widths = [14, 44, 14, 14, 18, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    info = wb.create_sheet("Instructions")
    lines = [
        ["Kosha import template — how to use"],
        [""],
        ["Put one transaction per row on the 'Transactions' sheet."],
        ["Columns:"],
        ["  Date", "required. e.g. 05/04/2026 or 2026-04-05"],
        ["  Transaction Remarks", "required. the narration / description"],
        ["  Debit", "amount that left the account (leave blank for credits)"],
        ["  Credit", "amount that came in (leave blank for debits)"],
        ["  Source", "optional. account/bank name, e.g. HDFC Bank, SBI Savings, Axis Card"],
        ["  Account Type", "optional. Bank or Credit Card (guessed from Source if blank)"],
        [""],
        ["Each row needs a Date and either a Debit or a Credit amount."],
        ["One file can mix multiple accounts — just set Source per row."],
        ["Example:"],
        ["  2026-04-05", "UPI SWIGGY", "300", "", "HDFC Bank", "Bank"],
        ["  2026-04-06", "Salary", "", "60000", "HDFC Bank", "Bank"],
    ]
    for line in lines:
        info.append(line)
    info["A1"].font = Font(bold=True, size=13)
    info.column_dimensions["A"].width = 24
    info.column_dimensions["B"].width = 52

    wb.save(str(path))
